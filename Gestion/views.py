from django.db.models.query import QuerySet
from django.forms import fields
from django.http.response import HttpResponse
from django.urls import reverse
from django.shortcuts import redirect, render,get_object_or_404,HttpResponseRedirect
from .models import Article, ArticleBL, Bon, Etablissement, Facture
from django.template import RequestContext, context
from django.contrib import messages
from .forms import ArticleBLForm, ArticleForm, BonForm, EtablissementForm, FactureForm
from django.forms import modelformset_factory
from django.forms import inlineformset_factory
from openpyxl import Workbook, load_workbook
import mimetypes
import os
from num2words import num2words
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side

def Accueil(request):
    return render(request,'Gestion/Accueil.html')

def Contact(request):
    return render(request,'Gestion/Contact.html') 

def Travail(request):
    return render(request, 'Gestion/Travail.html')

def GestionStock(request):
    articles = Article.objects.all()

    context = {
        'articles': articles ,
    }
    return render(request, 'Gestion/Gestion-du-Stock.html', context)

def SupprimerArticle(request,id):
    a = get_object_or_404(Article, pk=id)
    a.delete()
    messages.success(request, 'Vous avez supprimé un article avec succès')
    return redirect('GestionStock')

def ModifierArticle(request,id = id):
    context={}
    a = get_object_or_404(Article, pk=id)
    form = ArticleForm(request.POST or None, instance = a)
    if form.is_valid():
        form.save()
        messages.success(request, 'Vous avez modifié l\'article avec succès')
        return HttpResponseRedirect(request.path_info)
    context["form"] = form
    context["a"] = a
    return render(request, "Gestion/modifierArticle.html", context)

def AfficherArticle(request, id):
    context={}
    a = get_object_or_404(Article,id = id)
    context["a"] = a
    return render(request, "Gestion/consulterArticle.html", context)

def AjouterArticle(request):
    context={}
    form = ArticleForm(request.POST)
    context["form"] = form
    if form.is_valid():
        form.save()
        messages.success(request, 'Vous avez crée un article avec succès')
        return redirect('GestionStock')
    return render(request, "Gestion/ajouterArticle.html",context)

def ModifierStock(request):
    context={}
    articles=Article.objects.all()
    ArticleFormSet = modelformset_factory(Article, fields=("nbDisponible",), extra=0)
    if request.method == 'POST':
        formset = ArticleFormSet( request.POST) 
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Vous avez modifé le stock avec succès')
            return HttpResponseRedirect(request.path_info)
    else:
        formset = ArticleFormSet()
    context["ez"] = zip(articles,formset)
    context["formset"]=formset
    return render (request, "Gestion/ModifierStock.html",context )

    #-------------------------------Consultations----------------------------------------------------#

def Comptabilité(request):
    context={}
    e = Etablissement.objects.all()
    context['e']=e
    
    for etab in e:
        s=0
        for f in etab.facture_set.all():
            s = s + f.Montant
        etab.Montant = s
        etab.save()

    

    return render (request,"Gestion/Comptabilité.html",context )

def AjouterEtablissement(request):
    context={}
    form = EtablissementForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, 'Vous avez crée un établissement avec succès')
        return redirect('Comptabilité')
    context['form']=form
    return render(request, "Gestion/ajouterEtablissement.html",context)

def AjouterFacture(request):
    context={}
    form = FactureForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, 'Vous avez crée une Facture avec succès')
        return redirect('Comptabilité')
    context['form']=form
    return render(request, "Gestion/ajouterFacture.html",context)    


def AjouterBL(request):
    context={}
    form = BonForm(request.POST)
    if form.is_valid():
        obj = form.save()
        messages.success(request, 'Vous avez crée un Bon de Livraison avec succès')
        obj.save()
        return redirect('ConsulterBL', id = obj.id)
    context['form']=form
    return render(request, "Gestion/ajouterBL.html",context)

def ConsulterEtablissement(request,id):
    context={}
    e = get_object_or_404(Etablissement,id=id)
    context['e']=e
    dette = []
    for d in e.facture_set.all():
        dette.append(d.paye)
    context['dette']=dette

    return render(request, "Gestion/ConsulterEtablissement.html",context)

def ConsulterFacture(request,id):
    context={}
    s=0
    f = get_object_or_404(Facture,id=id)
    context['f']=f
    for b in f.Bons.all():
        s= s + b.Montant

    f.Montant = s
    f.save()
    context['s']=s
    return render(request, "Gestion/ConsulterFacture.html",context)

def ConsulterBL(request,id):
    context={}
    b = get_object_or_404(Bon,id=id)
    s=0
    context['b']=b
    form = ArticleBLForm(request.POST or None)
    articles = b.articlebl_set.all()


    if form.is_valid():
        obj=form.save()
        messages.success(request, 'Vous avez ajouté l\'article: '+obj.article.Nom +' avec succès')
        obj.Bon=b
        obj.save()


    for a in articles.all():
        s=s + (a.Quantité * a.article.prix)
    b.Montant = s
    b.save()
        
    context['Aform'] = form
    context['articles'] = articles

    return render(request, "Gestion/ConsulterBL.html",context)

#---------------------------------------Modifications-----------------#

def ModifierEtablissement(request,id):
    context={}
    e = get_object_or_404(Etablissement,id=id)
    form = EtablissementForm(request.POST, instance=e)
    if form.is_valid():
        form.save()
        messages.success(request, 'Vous avez modifé un établissement avec succès')
        return HttpResponseRedirect(request.path_info)
    context['form']=form
    context['e']=e
    return render(request, "Gestion/ModifierEtablissement.html",context)

def ModifierFacture(request,id):
    context={}
    f = get_object_or_404(Facture,id=id)
    form = FactureForm(request.POST or None, instance=f)
    if form.is_valid():
        form.save()
        messages.success(request, 'Vous avez modifé une facture avec succès')
        return HttpResponseRedirect(request.path_info)
    else:
        form = FactureForm(instance=f)

    context['form']=form
    context['f']=f
    return render(request, "Gestion/ModifierFacture.html",context)

def ModifierBL(request,id):
    context={}
    b = get_object_or_404(Bon,id=id)
    form = BonForm(request.POST or None, instance=b)
    if form.is_valid():
        form.save()
        messages.success(request, 'Vous avez modifé un bon de livraison avec succès')
        HttpResponseRedirect('Comptabilité')
    else:
        form = BonForm(instance=b)
        
    context['form']=form
    context['b']=b
    return render(request, "Gestion/ModifierBL.html",context)

#----------------------------------Suppressions---------------------------------#
def SupprimerEtablissement(request,id):
    e = get_object_or_404(Etablissement, pk=id)
    e.delete()
    messages.success(request, 'Vous avez supprimé un établissement avec succès')
    return redirect('Comptabilité')

def SupprimerFacture(request,id):
    f = get_object_or_404(Facture, pk=id)
    f.delete()
    messages.success(request, 'Vous avez supprimé une facture avec succès')
    return redirect('Comptabilité')

def SupprimerBL(request,id):
    b = get_object_or_404(Bon, pk=id)
    b.delete()
    messages.success(request, 'Vous avez supprimé un bon de livraison avec succès')
    return redirect('Comptabilité')

def SupprimerArticleBL(request,id):
    b = get_object_or_404(ArticleBL, pk=id)
    b.delete()
    messages.success(request, 'Vous avez supprimé un article avec succès')
    return redirect('ConsulterBL',b.Bon.id)

#-----------------------------------------------------------------------------------------------------#


def ExportBL(request,id):
    
    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    img = Image('Aitsaid.png')
    img.anchor = 'B1'
    img.height = 50
    img.width = 250

    B = Bon.objects.get(id=id)
    wb = load_workbook(filename='BL_template.xlsx', keep_vba=True)
    w1 = wb.active
    
    w1['B19'] = B.Etab.Nom
    w1['E16'] = "N°"+str(B.Num)
    w1['E17'] = str(B.date)
    w1['B25'] = (num2words(B.Montant, lang='fr')).upper() +" DINARS"
    w1.add_image(img, 'B1')
    data=[]
    count=0
    step = 22
    for a in B.articlebl_set.all(): 
        count +=1
        data.append([count, (a.article.Nom + " "+ a.article.Marque), str(a.Quantité), str("{:10.2f}".format(a.article.prix)), str("{:10.2f}".format(a.Quantité * a.article.prix)) ]) 
    
    for d in data:
        w1.insert_rows(step)
        w1.cell(row=step, column=1, value= d[0]) 
        w1.cell(row=step, column=2, value= d[1]) 
        w1.cell(row=step, column=3, value= d[2]) 
        w1.cell(row=step, column=4, value= d[3]) 
        w1.cell(row=step, column=5, value= d[4]) 
        step = step +1
    
    set_border(w1, 'A22:E'+str(step))
    w1.cell(row=step+1, column=5, value= B.Montant) 
    


    filepath= B.__str__() +".xls"

    wb.save(filename = filepath)
    path = open(filepath, 'rb')
    mime_type, _ = mimetypes.guess_type(filepath)
    response = HttpResponse(path, content_type=mime_type)
    response['Content-Disposition'] = "attachment; filename=%s" % filepath
    return response


def ExportFac(request, id):
    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    img = Image('Aitsaid.png')
    img.anchor = 'B1'
    img.height = 50
    img.width = 250

    F = Facture.objects.get(id=id)
    wb = load_workbook(filename='Fac_template.xlsx', keep_vba=True)
    w1 = wb.active
    
    w1['B19'] = F.Etab.Nom
    w1['E16'] = "N°"+str(F.Num)
    w1['E17'] = str(F.date)
    w1['B25'] = (num2words(F.Montant, lang='fr')).upper() +" DINARS"
    w1.add_image(img, 'B1')
    data=[]
    count=0
    step = 22
    for b in F.Bons.all(): 
        for a in b.articlebl_set.all():
            count +=1
            data.append([count, (a.article.Nom + " "+ a.article.Marque), str(a.Quantité), str("{:10.2f}".format(a.article.prix)), str("{:10.2f}".format(a.Quantité * a.article.prix)) ]) 
    

    for d in data:
        w1.insert_rows(step)
        w1.cell(row=step, column=1, value= d[0]) 
        w1.cell(row=step, column=2, value= d[1]) 
        w1.cell(row=step, column=3, value= d[2]) 
        w1.cell(row=step, column=4, value= d[3]) 
        w1.cell(row=step, column=5, value= d[4]) 
        step = step +1
    
    set_border(w1, 'A22:E'+str(step))
    w1.cell(row=step+1, column=5, value= F.Montant) 
    


    filepath= F.__str__() +".xls"

    wb.save(filename = filepath)
    path = open(filepath, 'rb')
    mime_type, _ = mimetypes.guess_type(filepath)
    response = HttpResponse(path, content_type=mime_type)
    response['Content-Disposition'] = "attachment; filename=%s" % filepath
    return response







